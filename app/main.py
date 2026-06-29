"""
Tidal Access Window Predictor - FastAPI application.

Single-page web app with API routes for configuration, calculation,
data management, and iCal feed serving.
"""

import logging
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, Request, HTTPException, Depends
from fastapi.responses import HTMLResponse, Response, JSONResponse
from fastapi.staticfiles import StaticFiles
from dateutil import parser as dtparse

from app.config import (
    ensure_dirs, load_model_config,
    FEEDS_DIR, DEFAULT_TIMEZONE, UKHO_STATION_ID, OWM_API_KEY,
    LOCATION_LAT, LOCATION_LON,
    to_utc_str,
)
from app.database import (
    init_db, save_mooring, get_mooring, get_all_moorings, delete_mooring,
    add_observation, get_observations, delete_observation, clear_observations,
    store_tide_events, get_tide_events, get_ukho_tide_events,
    calibrate_drying_height, calibrate_wind_offset,
    load_classification_inputs,
    delete_future_events, get_calendar_events, log_activity, get_activity_log,
    get_mooring_pin_hash, set_mooring_pin_hash,
    check_pin_lockout, record_failed_pin_attempt, clear_failed_pin_attempts,
    get_harmonic_predictions,
)
from app.pin import (
    hash_pin, verify_pin, is_valid_pin_format,
    MAX_PIN_ATTEMPTS, PIN_ATTEMPT_WINDOW_MINUTES, PIN_LOCKOUT_MINUTES,
)
from app.ukho import fetch_tidal_events
from app.harmonic import predict_events as harmonic_predict_events
from app.secondary_port import apply_offset
from app.wind import fetch_current_wind, fetch_current_weather
from app.conditions import get_current_conditions
from app.access_calc import compute_access_windows, generate_event_uid, _interpolate_from_parsed
from app.ical_manager import (
    generate_export_ics, store_windows_as_events,
    generate_feed_for_mooring,
    generate_langstone_ukho_7d_feed, generate_langstone_harmonic_180d_feed,
    generate_langstone_ukho_7d_pressure_corrected_feed,
)
from app.scheduler import start_scheduler, shutdown_scheduler, ensure_wind_jobs_scheduled

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
)
logger = logging.getLogger(__name__)


def _coalesce(*values, default):
    """
    Return the first value in *values that is not None, otherwise default.

    Purpose: replaces truthiness-based `a or b` idiom where a legitimate
    zero/empty/False value must not be overridden by a fallback. For example,
    a drying_height_m of 0.0 or a negative value must survive unchanged.

    Usage: _coalesce(data.get("x"), mooring["x"] if mooring else None, default=1.0)
    """
    for v in values:
        if v is not None:
            return v
    return default


def _public_mooring(m):
    """
    Return a copy of a mooring dict with the sensitive pin_hash field
    stripped out. Use this when returning mooring data to API clients.
    A None input returns None so call sites can use this uniformly.
    """
    if m is None:
        return None
    return {k: v for k, v in m.items() if k != "pin_hash"}


def _verify_pin_from_request(
    mooring_id: int,
    request: Request,
    allow_unclaimed: bool = False,
    allow_nonexistent: bool = False,
) -> None:
    """
    Verify the X-Mooring-PIN header against the stored hash for the given
    mooring. Raises HTTPException on any failure path; returns None on
    success.

    Failure responses (as HTTPException with structured detail):
      - 429 if the mooring is currently locked out. Sends Retry-After.
      - 404 if the mooring does not exist and allow_nonexistent is False.
      - 403 with pin_required=claim if the mooring exists but has no PIN
        set and allow_unclaimed is False. The UI interprets this as a
        prompt to call POST /pin to claim.
      - 401 with pin_required=verify if the header is missing.
      - 401 with pin_required=verify and attempts_remaining on PIN
        mismatch. The failed-attempt counter is incremented.
      - 429 if the wrong-PIN attempt is the one that trips the lockout.

    Side effects:
      - On success: clears any stored failed-attempt counter.
      - On PIN mismatch: increments the failed-attempt counter and may
        start a lockout window per the policy in app.pin.

    The two allow_* flags exist to support POST /api/moorings:
      - allow_nonexistent=True lets the first-ever save on a fresh
        mooring_id pass through (there is no row yet to read a hash from).
      - allow_unclaimed=True lets a save on an existing-but-unclaimed
        mooring pass through (the UI will follow up with a claim call).
    All other callers use the strict defaults.
    """
    lockout = check_pin_lockout(mooring_id)
    if lockout:
        raise HTTPException(
            status_code=429,
            detail={
                "message": "Too many failed PIN attempts",
                "locked_until": lockout["locked_until"],
                "seconds_remaining": lockout["seconds_remaining"],
            },
            headers={"Retry-After": str(lockout["seconds_remaining"])},
        )

    stored_hash = get_mooring_pin_hash(mooring_id)
    if stored_hash is None:
        if allow_nonexistent:
            return
        raise HTTPException(404, "Mooring not found")

    if stored_hash == "":
        if allow_unclaimed:
            return
        raise HTTPException(
            status_code=403,
            detail={
                "message": "Mooring has no PIN set",
                "pin_required": "claim",
            },
        )

    pin = request.headers.get("X-Mooring-PIN")
    if not pin:
        raise HTTPException(
            status_code=401,
            detail={
                "message": "PIN required",
                "pin_required": "verify",
            },
        )

    if verify_pin(pin, stored_hash):
        clear_failed_pin_attempts(mooring_id)
        return

    state = record_failed_pin_attempt(
        mooring_id,
        MAX_PIN_ATTEMPTS,
        PIN_ATTEMPT_WINDOW_MINUTES,
        PIN_LOCKOUT_MINUTES,
    )
    if state["locked_until"]:
        raise HTTPException(
            status_code=429,
            detail={
                "message": "Too many failed attempts; locked out",
                "locked_until": state["locked_until"],
                "attempts_remaining": 0,
            },
            headers={"Retry-After": str(PIN_LOCKOUT_MINUTES * 60)},
        )
    raise HTTPException(
        status_code=401,
        detail={
            "message": "Incorrect PIN",
            "pin_required": "verify",
            "attempts_remaining": state["attempts_remaining"],
        },
    )


async def require_mooring_pin(mooring_id: int, request: Request) -> None:
    """
    FastAPI dependency wrapping _verify_pin_from_request with the strict
    defaults: rejects non-existent moorings (404) and unclaimed moorings
    (403 with pin_required=claim). Use via
    `dependencies=[Depends(require_mooring_pin)]` on write endpoints
    where the mooring must already exist and already be claimed.
    """
    _verify_pin_from_request(
        mooring_id, request, allow_unclaimed=False, allow_nonexistent=False
    )


def _build_calibration_response(mooring_id: int) -> dict:
    """
    Build the combined calibration response: base drying fields at top
    level (for back-compat), wind_offset sub-object, and a classifications
    map keyed by observation id. Used by GET /calibration and as the
    response body for observation mutations.

    All three components (drying calibration, wind offset calibration,
    classifications map) share a single data-loading pass via
    load_classification_inputs, avoiding redundant DB queries and
    classifier passes.
    """
    preloaded = load_classification_inputs(mooring_id)
    base = calibrate_drying_height(mooring_id, _preloaded=preloaded)
    wind = calibrate_wind_offset(mooring_id, _preloaded=preloaded)

    # Build classifications map from the pre-loaded classifier output
    classifications_map = {}
    _mooring, _observations, _tide_events, _wind_obs, classified = preloaded
    for entry in classified:
        obs_id = entry["observation"].get("id")
        if obs_id is None:
            continue
        classifications_map[str(obs_id)] = {
            "classification": entry["classification"],
            "reason": entry["reason"],
            "hw_timestamp": entry["hw_timestamp"],
            "wind_compass": entry["wind_compass"],
        }

    response = dict(base)
    response["wind_offset"] = wind
    response["classifications"] = classifications_map
    return response


def _compute_tender_windows(
    events: list[dict],
    mooring: dict,
    source: str,
) -> tuple[Optional[list[dict]], Optional[float]]:
    """
    Run a second compute_access_windows() pass with draught_m=0 and
    safety_margin_m=tender_min_depth_m so the threshold becomes
    drying_height_m + tender_min_depth_m. Returns (windows, depth) when
    tender access is enabled on the mooring, else (None, None).

    These are baseline (no wind) windows. Wind adjustment is applied
    separately by the scheduler via compute_next_window_with_wind, which
    runs its own tender pass with the same offset.
    """
    if not mooring or not mooring.get("tender_access_enabled"):
        return None, None
    depth = float(mooring.get("tender_min_depth_m") or 0.3)
    tw = compute_access_windows(
        events=events,
        draught_m=0.0,
        drying_height_m=mooring["drying_height_m"],
        safety_margin_m=depth,
        source=source,
    )
    return tw, depth


def _recompute_future_windows(mooring_id: int):
    """
    After a mooring configuration change, clear stored future events and
    recompute them from currently-stored UKHO tide data, then regenerate
    the iCal feed if calendar subscription is enabled.

    Does nothing if no UKHO data is available. Uses the mooring's current
    stored configuration. Uses get_ukho_tide_events() so that Portsmouth
    fallback data receives the secondary port correction transparently.
    """
    m = get_mooring(mooring_id)
    if not m:
        return

    now = datetime.now(timezone.utc)
    now_str = to_utc_str(now)
    delete_future_events(mooring_id, now_str)

    query_start = to_utc_str(now - timedelta(hours=13))
    end = to_utc_str(now + timedelta(days=7))
    tide_data = get_ukho_tide_events(query_start, end)
    if not tide_data:
        return

    # Barometric correction (v2.9): calibration-apply rebuilds this mooring's
    # future windows from scratch (delete_future_events above), so correct the
    # event heights for an opted-in mooring before recomputing -- otherwise the
    # rebuilt feed would briefly drop the pressure correction (overstating
    # access under high pressure) until the next daily job. Gated on the system
    # master AND opt-in, so it is a no-op while the feature is dark. No deadband:
    # the rows were just deleted, so this is a deliberate full rebuild.
    from app.config import get_barometric_enabled
    if get_barometric_enabled(False) and m.get("barometric_enabled"):
        from app.barometric import apply_barometric_correction, make_pressure_provider
        tide_data = apply_barometric_correction(tide_data, make_pressure_provider())

    windows = compute_access_windows(
        events=tide_data,
        draught_m=m["draught_m"],
        drying_height_m=m["drying_height_m"],
        safety_margin_m=m["safety_margin_m"],
        source="ukho",
    )
    # Filter out windows for HW events in the 13h lookback window — those
    # were included only for interpolation context, not as real future windows.
    windows = [w for w in windows if w["hw_timestamp"] >= now_str]

    tender_windows, tender_depth = _compute_tender_windows(tide_data, m, "ukho")

    cal = calibrate_drying_height(mooring_id)
    calc_params = {
        "draught_m": m["draught_m"],
        "drying_height_m": m["drying_height_m"],
        "safety_margin_m": m["safety_margin_m"],
        "obs_calibrated": 1 if cal.get("confidence", "none") != "none" else 0,
    }
    store_windows_as_events(
        windows, mooring_id, "ukho", m.get("boat_name", ""),
        calc_params=calc_params,
        tender_windows=tender_windows,
        tender_min_depth_m=tender_depth,
    )

    if m.get("calendar_enabled"):
        feed_cal = cal if cal.get("confidence") != "none" else None
        generate_feed_for_mooring(mooring_id, m.get("boat_name", ""), feed_cal)

    # Config/calibration changed -> reschedule wind jobs so the next grounding
    # samples against the new draught/drying/offset. Idempotent global rebuild.
    try:
        ensure_wind_jobs_scheduled()
    except Exception as e:
        logger.warning(f"Wind-job reschedule after recompute failed: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application startup and shutdown."""
    ensure_dirs()
    init_db()
    load_model_config()  # Loads bundled config and warms the cache via _get_curve_params on first calc
    _warm_up_harmonic_predictions()
    start_scheduler()
    # Rebuild per-mooring wind-sample jobs immediately on boot. The in-memory
    # APScheduler store is empty after a restart, so without this the jobs would
    # not exist until the next 02:00 daily fetch (the 15-min conditions refresh
    # is a slower safety net).
    try:
        ensure_wind_jobs_scheduled()
    except Exception as e:
        logger.warning(f"Initial wind-job scheduling failed: {e}")
    logger.info("Tidal Access application started")
    yield
    shutdown_scheduler()
    logger.info("Tidal Access application stopped")


def _warm_up_harmonic_predictions():
    """
    On a fresh deployment the harmonic_predictions table is empty until
    the first 02:00 scheduler run. Pre-populate it with 180 days of
    Langstone-corrected predictions so the tide-curve panel can render
    harmonic-source days (today+7..today+180) immediately.

    Idempotent: only runs when the table is empty. Subsequent restarts
    are no-ops because the daily scheduler keeps the table populated.
    Synchronous and fast (~1-3s) for a clean process start.
    """
    try:
        from app.database import db_connection, store_harmonic_predictions
        from app.harmonic import predict_events as harmonic_predict_events
        with db_connection() as conn:
            count = conn.execute("SELECT COUNT(*) FROM harmonic_predictions").fetchone()[0]
        if count > 0:
            logger.info(f"Harmonic predictions table populated ({count} rows); skipping warm-up.")
            return
        logger.info("Harmonic predictions table empty; generating 180-day warm-up batch...")
        start = datetime.now(timezone.utc)
        end = start + timedelta(days=180)
        raw = harmonic_predict_events(start, end)
        langstone = apply_offset(raw) if raw else []
        inserted = store_harmonic_predictions(langstone)
        logger.info(f"Harmonic warm-up complete: {inserted} predictions stored.")
        log_activity(
            event_type="harmonic_refresh",
            message=f"Startup warm-up: stored {inserted} harmonic predictions",
            severity="success" if inserted else "warning",
            details={"event_count": inserted, "window_days": 180, "trigger": "startup_warmup"},
        )
    except Exception as e:
        logger.warning(f"Harmonic startup warm-up failed: {e}")


app = FastAPI(title="Tidal Access Window Predictor", lifespan=lifespan)

# Serve static files
STATIC_DIR = Path(__file__).parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# --- Pages ---

@app.get("/", response_class=HTMLResponse)
async def index():
    """Serve the single-page web UI."""
    html_path = STATIC_DIR / "index.html"
    return HTMLResponse(html_path.read_text())


# --- Mooring Configuration ---

@app.get("/api/moorings")
async def list_moorings():
    """List all configured moorings with sensitive fields (pin_hash) stripped."""
    return [_public_mooring(m) for m in get_all_moorings()]


@app.get("/api/moorings/{mooring_id}")
async def get_mooring_config(mooring_id: int):
    """Get a specific mooring configuration with pin_hash stripped."""
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")
    return _public_mooring(m)


@app.post("/api/moorings")
async def save_mooring_config(request: Request):
    """
    Save (create or update) a mooring configuration.

    PIN gating: Existing moorings with a non-empty pin_hash require a
    valid X-Mooring-PIN header. New moorings (no row yet) and existing
    moorings with no PIN set (unclaimed) are allowed through; the UI
    should follow up with a call to POST /api/moorings/{id}/pin to
    claim the PIN.
    """
    data = await request.json()
    if "mooring_id" not in data:
        raise HTTPException(400, "mooring_id is required")
    mid = int(data["mooring_id"])
    if mid < 1 or mid > 100:
        raise HTTPException(400, "mooring_id must be between 1 and 100")

    # PIN gate for claimed moorings only. Brand-new mooring_ids and
    # unclaimed existing ones pass through so the UI can immediately
    # prompt to claim via POST /pin.
    _verify_pin_from_request(
        mid, request, allow_unclaimed=True, allow_nonexistent=True
    )

    result = save_mooring(data)

    # Reschedule wind-sample jobs so changes to wind/shallow or draught/drying
    # settings take effect now rather than waiting for the next daily fetch.
    try:
        ensure_wind_jobs_scheduled()
    except Exception as e:
        logger.warning(f"Wind-job reschedule after config save failed: {e}")

    log_activity(
        event_type="mooring_config",
        message=(
            f"Configuration saved for mooring #{mid}"
            + (f" ({result.get('boat_name')})" if result.get("boat_name") else "")
        ),
        severity="info",
        scope="mooring",
        mooring_id=mid,
        details={
            "boat_name": result.get("boat_name", ""),
            "draught_m": result.get("draught_m"),
            "drying_height_m": result.get("drying_height_m"),
            "safety_margin_m": result.get("safety_margin_m"),
            "calendar_enabled": bool(result.get("calendar_enabled")),
            "wind_offset_enabled": bool(result.get("wind_offset_enabled")),
            "barometric_enabled": bool(result.get("barometric_enabled")),
            "shallow_direction": result.get("shallow_direction", ""),
            "shallow_extra_depth_m": result.get("shallow_extra_depth_m", 0),
            "use_observations": bool(result.get("use_observations")),
        },
    )

    if result.get("calendar_enabled"):
        existing_events = get_calendar_events(mid)
        if existing_events:
            cal = calibrate_drying_height(mid)
            if cal.get("confidence") == "none":
                cal = None
            generate_feed_for_mooring(mid, result.get("boat_name", ""), cal)

    return _public_mooring(result)


@app.post("/api/moorings/{mooring_id}/pin")
async def set_or_change_pin(mooring_id: int, request: Request):
    """
    Claim or change the PIN for a mooring.

    For unclaimed moorings (no PIN set): only new_pin is required.
    For claimed moorings: both current_pin (for verification) and new_pin
    are required. The same rate-limit policy applies as for other PIN
    operations - a wrong current_pin counts toward the lockout.

    Body: {"new_pin": "123456", "current_pin": "654321" (optional)}
    Returns: {"status": "ok", "was_claimed": bool}
    """
    data = await request.json()
    new_pin = data.get("new_pin", "") or ""
    current_pin = data.get("current_pin", "") or ""

    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")

    if not is_valid_pin_format(new_pin):
        raise HTTPException(400, "PIN must be exactly six numeric digits")

    # Respect the lockout timer for any PIN operation.
    lockout = check_pin_lockout(mooring_id)
    if lockout:
        raise HTTPException(
            status_code=429,
            detail={
                "message": "Too many failed PIN attempts",
                "locked_until": lockout["locked_until"],
                "seconds_remaining": lockout["seconds_remaining"],
            },
            headers={"Retry-After": str(lockout["seconds_remaining"])},
        )

    stored_hash = get_mooring_pin_hash(mooring_id) or ""
    was_claimed = bool(stored_hash)

    if was_claimed:
        if not current_pin:
            raise HTTPException(
                status_code=401,
                detail={
                    "message": "current_pin is required to change an existing PIN",
                    "pin_required": "verify",
                },
            )
        if not verify_pin(current_pin, stored_hash):
            state = record_failed_pin_attempt(
                mooring_id,
                MAX_PIN_ATTEMPTS,
                PIN_ATTEMPT_WINDOW_MINUTES,
                PIN_LOCKOUT_MINUTES,
            )
            if state["locked_until"]:
                raise HTTPException(
                    status_code=429,
                    detail={
                        "message": "Too many failed attempts; locked out",
                        "locked_until": state["locked_until"],
                        "attempts_remaining": 0,
                    },
                    headers={"Retry-After": str(PIN_LOCKOUT_MINUTES * 60)},
                )
            raise HTTPException(
                status_code=401,
                detail={
                    "message": "Incorrect current PIN",
                    "pin_required": "verify",
                    "attempts_remaining": state["attempts_remaining"],
                },
            )
        clear_failed_pin_attempts(mooring_id)

    new_hash = hash_pin(new_pin)
    if not set_mooring_pin_hash(mooring_id, new_hash):
        # Should not happen - we just confirmed the row exists above.
        raise HTTPException(500, "Failed to update PIN")

    log_activity(
        event_type="pin_changed" if was_claimed else "pin_claimed",
        message=(
            f"PIN changed for mooring #{mooring_id}"
            if was_claimed
            else f"PIN claimed for mooring #{mooring_id}"
        ),
        severity="success",
        scope="mooring",
        mooring_id=mooring_id,
    )

    return {"status": "ok", "was_claimed": was_claimed}


@app.delete("/api/moorings/{mooring_id}",
            dependencies=[Depends(require_mooring_pin)])
async def delete_mooring_config(mooring_id: int):
    """
    Permanently delete a mooring and its associated per-mooring data.
    PIN-gated. Returns 404 if the mooring does not exist (the dependency
    handles this before the handler runs).

    Database state cleared: moorings row, observations, calendar_events,
    pin_failed_attempts. Activity log entries are preserved as an audit
    trail (and age out automatically per prune_activity_log retention).

    On-disk state cleared: feeds/mooring_NN.ics if present.

    Returns a summary dict with deletion counts.
    """
    m = get_mooring(mooring_id)
    boat_name = m.get("boat_name") if m else ""

    counts = delete_mooring(mooring_id)

    # Remove the on-disk feed file. Filename pattern matches
    # generate_feed_for_mooring (zero-padded to 3 digits, hence :03d).
    # If the file does not exist, this is a no-op - users may have
    # never enabled calendar subscription.
    feed_path = FEEDS_DIR / f"mooring_{mooring_id:03d}.ics"
    feed_removed = False
    try:
        if feed_path.exists():
            feed_path.unlink()
            feed_removed = True
    except OSError as e:
        # Log but do not fail the deletion - the database row is gone, the
        # feed file is now orphaned, and serve_feed will 404 since the
        # mooring no longer exists. Manual cleanup of the file is possible
        # if needed.
        logger.warning(
            f"Failed to remove feed file {feed_path} during mooring delete: {e}"
        )

    log_activity(
        event_type="mooring_deleted",
        message=(
            f"Mooring #{mooring_id}"
            + (f" ({boat_name})" if boat_name else "")
            + " permanently deleted"
        ),
        severity="warning",
        scope="mooring",
        mooring_id=mooring_id,
        details={
            "boat_name": boat_name,
            "db_counts": counts,
            "feed_file_removed": feed_removed,
        },
    )

    return {
        "deleted": True,
        "mooring_id": mooring_id,
        "db_counts": counts,
        "feed_file_removed": feed_removed,
    }


# --- Observations ---

@app.get("/api/moorings/{mooring_id}/observations")
async def list_observations(mooring_id: int):
    """Get all observations for a mooring."""
    return get_observations(mooring_id)


@app.delete("/api/moorings/{mooring_id}/observations/{observation_id}",
            dependencies=[Depends(require_mooring_pin)])
async def remove_observation(mooring_id: int, observation_id: int):
    """
    Delete a single observation. Does NOT auto-apply calibration changes -
    the updated calibration is returned for the UI to display as a
    suggestion. The user applies it explicitly via the Apply endpoints.
    """
    if not delete_observation(observation_id, mooring_id):
        raise HTTPException(404, "Observation not found")

    log_activity(
        event_type="observation_deleted",
        message=f"Observation #{observation_id} removed",
        severity="info",
        scope="mooring",
        mooring_id=mooring_id,
    )

    return {
        "deleted": observation_id,
        "calibration": _build_calibration_response(mooring_id),
    }


@app.delete("/api/moorings/{mooring_id}/observations",
            dependencies=[Depends(require_mooring_pin)])
async def clear_all_observations(mooring_id: int):
    """
    Delete all observations for a mooring. Does NOT auto-apply calibration
    changes - the mooring's stored drying height and shallow_extra_depth
    are left untouched; the UI can prompt the user to re-set them manually.
    """
    count = clear_observations(mooring_id)
    if count > 0:
        log_activity(
            event_type="observations_cleared",
            message=f"All {count} observations cleared",
            severity="warning",
            scope="mooring",
            mooring_id=mooring_id,
            details={"count": count},
        )
    return {
        "deleted_count": count,
        "calibration": _build_calibration_response(mooring_id),
    }


@app.post("/api/moorings/{mooring_id}/observations",
          dependencies=[Depends(require_mooring_pin)])
async def add_mooring_observation(mooring_id: int, request: Request):
    """
    Add an observation. Does NOT auto-apply calibration changes; the
    returned calibration payload is a suggestion for the UI to render
    with Apply buttons.
    """
    data = await request.json()
    data["mooring_id"] = mooring_id

    # v2.10: a sounding records a raw measured depth rather than an
    # afloat/aground state. Validate the depth and datum up front so a
    # malformed reading is rejected rather than silently dropped at
    # derivation time. Binary observations are unaffected.
    if (data.get("obs_type") or "binary") == "sounding":
        from app.access_calc import SOUNDER_DATUMS
        try:
            depth = float(data.get("measured_depth_m"))
        except (TypeError, ValueError):
            raise HTTPException(400, "Sounding requires a numeric measured_depth_m")
        if depth < 0:
            raise HTTPException(400, "measured_depth_m must be non-negative")
        datum = (data.get("sounder_datum") or "").strip().lower()
        if datum and datum not in SOUNDER_DATUMS:
            raise HTTPException(
                400, f"sounder_datum must be one of {', '.join(SOUNDER_DATUMS)}"
            )

    # Existence check is handled by the require_mooring_pin dependency,
    # which returns 404 before this handler runs.
    obs = add_observation(data)

    is_sounding = obs.get("obs_type") == "sounding"
    if is_sounding:
        msg = (f"Sounding added: {data.get('measured_depth_m')}m "
               f"({data.get('sounder_datum') or 'boat default'}) at "
               f"{obs.get('timestamp', '')[:16]}")
    else:
        msg = f"Observation added: {obs.get('state')} at {obs.get('timestamp', '')[:16]}"
    log_activity(
        event_type="observation_added",
        message=msg,
        severity="info",
        scope="mooring",
        mooring_id=mooring_id,
        details={
            "obs_type": obs.get("obs_type", "binary"),
            "state": obs.get("state"),
            "measured_depth_m": data.get("measured_depth_m"),
            "sounder_datum": data.get("sounder_datum"),
            "timestamp": obs.get("timestamp"),
            "wind_direction": obs.get("wind_direction", ""),
            "direction_of_lay": obs.get("direction_of_lay", ""),
            "source": "manual",
        },
    )

    return {
        "observation": obs,
        "calibration": _build_calibration_response(mooring_id),
    }


@app.post("/api/moorings/{mooring_id}/observations/upload",
          dependencies=[Depends(require_mooring_pin)])
async def upload_observations_xlsx(mooring_id: int, request: Request):
    """
    Import observations from an XLSX file. Observations are tied to this
    mooring only. Expected columns: Date, Time, State, Wind Direction,
    Direction of Lay, Notes.

    Does NOT auto-apply calibration - the returned calibration payload
    is a suggestion.
    """
    import io
    from openpyxl import load_workbook

    # Existence check is handled by the require_mooring_pin dependency.
    body = await request.body()
    if not body:
        raise HTTPException(400, "No file uploaded")

    try:
        wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Could not read XLSX file: {e}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    imported = 0
    errors = 0

    for row in rows:
        if not row or len(row) < 3:
            errors += 1
            continue

        date_val, time_val, state_val = row[0], row[1], row[2]
        wind_dir = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        lay_dir = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        notes = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        # v2.10 optional columns. Absent in pre-v2.10 templates, so a missing
        # or blank Obs Type is treated as a binary afloat/aground row exactly
        # as before.
        obs_type = str(row[6]).strip().lower() if len(row) > 6 and row[6] else "binary"
        depth_val = row[7] if len(row) > 7 else None
        datum_val = str(row[8]).strip().lower() if len(row) > 8 and row[8] else ""

        try:
            if isinstance(date_val, datetime):
                dt = date_val
            elif isinstance(date_val, str):
                dt = dtparse.parse(date_val)
            else:
                errors += 1
                continue

            if time_val is not None and not isinstance(date_val, datetime):
                time_str = str(time_val).strip()
                if ":" in time_str:
                    parts = time_str.split(":")
                    dt = dt.replace(hour=int(parts[0]), minute=int(parts[1]))
        except (ValueError, TypeError):
            errors += 1
            continue

        if dt.tzinfo is None:
            import pytz
            local_tz = pytz.timezone(DEFAULT_TIMEZONE)
            dt = local_tz.localize(dt)
        ts = to_utc_str(dt)

        if obs_type == "sounding":
            try:
                depth = float(depth_val)
            except (TypeError, ValueError):
                errors += 1
                continue
            add_observation({
                "mooring_id": mooring_id,
                "timestamp": ts,
                "obs_type": "sounding",
                "measured_depth_m": depth,
                "sounder_datum": datum_val or None,
                "wind_direction": wind_dir,
                "direction_of_lay": lay_dir,
                "notes": notes,
            })
            imported += 1
            continue

        state_str = str(state_val).strip().lower()
        if state_str in ("afloat", "yes", "y", "true"):
            state = "afloat"
        elif state_str in ("aground", "no", "n", "false"):
            state = "aground"
        else:
            errors += 1
            continue

        add_observation({
            "mooring_id": mooring_id,
            "timestamp": ts,
            "state": state,
            "wind_direction": wind_dir,
            "direction_of_lay": lay_dir,
            "notes": notes,
        })
        imported += 1

    if imported > 0:
        log_activity(
            event_type="observations_uploaded",
            message=f"Imported {imported} observations from XLSX ({errors} rows skipped)",
            severity="info",
            scope="mooring",
            mooring_id=mooring_id,
            details={"imported": imported, "errors": errors},
        )

    return {
        "imported": imported,
        "errors": errors,
        "calibration": _build_calibration_response(mooring_id),
    }


@app.get("/api/observations/template")
async def download_observation_template():
    """Generate and serve an XLSX template for observation recording."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "Observations"

    headers = ["Date", "Time", "State", "Wind Direction", "Direction of Lay", "Notes",
               "Obs Type", "Measured Depth (m)", "Sounder Datum"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example binary (afloat/aground) row — leave the v2.10 columns blank.
    ws.cell(row=2, column=1, value="15/04/2026")
    ws.cell(row=2, column=2, value="10:30")
    ws.cell(row=2, column=3, value="afloat")
    ws.cell(row=2, column=4, value="SW")
    ws.cell(row=2, column=5, value="NE")
    ws.cell(row=2, column=6, value="Spring tide, good visibility")

    # Example sounding row — State is left blank; Obs Type drives it.
    ws.cell(row=3, column=1, value="15/04/2026")
    ws.cell(row=3, column=2, value="11:15")
    ws.cell(row=3, column=6, value="Depth sounding at mooring")
    ws.cell(row=3, column=7, value="sounding")
    ws.cell(row=3, column=8, value=2.4)
    ws.cell(row=3, column=9, value="transducer")

    for col_letter, width in (
        ("A", 14), ("B", 8), ("C", 10), ("D", 16), ("E", 18), ("F", 30),
        ("G", 12), ("H", 18), ("I", 14),
    ):
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("Notes")
    ws2.cell(row=1, column=1, value="State: 'afloat' or 'aground' (binary rows only)")
    ws2.cell(row=2, column=1, value="Wind Direction: N, NE, E, SE, S, SW, W, NW (optional)")
    ws2.cell(row=3, column=1, value="Direction of Lay: bow heading N, NE, E, SE, S, SW, W, NW (optional)")
    ws2.cell(row=4, column=1, value="Date format: DD/MM/YYYY")
    ws2.cell(row=5, column=1, value="Time format: HH:MM (local time)")
    ws2.cell(row=6, column=1, value="Times are assumed to be local time (BST during sailing season)")
    ws2.cell(row=7, column=1, value="Obs Type: leave blank (or 'binary') for afloat/aground; 'sounding' for a depth reading")
    ws2.cell(row=8, column=1, value="Measured Depth (m): raw echo-sounder reading; required for sounding rows")
    ws2.cell(row=9, column=1, value="Sounder Datum: waterline, transducer or keel (blank = boat default)")
    ws2.cell(row=10, column=1, value="Soundings: State is ignored; the raw depth is converted to drying height at calibration time")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=observation_template.xlsx"},
    )


@app.get("/api/moorings/{mooring_id}/calibration")
async def get_calibration_status(mooring_id: int):
    """
    Get calibration status for a mooring.

    Response shape:
      - Base drying fields at the top level: best_estimate, lower_bound,
        upper_bound, confidence, matched, unmatched, afloat_count,
        aground_count, excluded_wind_offset_count.
      - `wind_offset`: suggested shallow-side offset with confidence and
        current stored values.
      - `classifications`: map from observation id (string) to its
        classifier output (classification, reason, hw_timestamp,
        wind_compass).
    """
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")
    return _build_calibration_response(mooring_id)


@app.post("/api/moorings/{mooring_id}/calibration/apply-drying-height",
          dependencies=[Depends(require_mooring_pin)])
async def apply_drying_height_calibration(mooring_id: int):
    """
    Apply the current base-drying-height suggestion to the mooring's
    stored configuration, then recompute future access windows and
    regenerate the feed.
    """
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")

    cal = calibrate_drying_height(mooring_id)
    best = cal.get("best_estimate")
    if best is None:
        raise HTTPException(400, "No drying height suggestion available to apply")
    if cal.get("confidence") == "inconsistent":
        # Aground observation lies above an afloat one — the midpoint is
        # arithmetic, not physical. Refuse to apply silently; user must
        # revisit the observations first.
        raise HTTPException(
            400,
            "Calibration bounds are inconsistent (aground height above an "
            "afloat height). Resolve the conflicting observation(s) before "
            "applying.",
        )

    previous = m["drying_height_m"]
    if abs(best - previous) < 0.01:
        return {
            "applied": False,
            "reason": "suggestion matches current stored value",
            "previous": previous,
            "new": previous,
            "mooring": _public_mooring(m),
            "calibration": _build_calibration_response(mooring_id),
        }

    m["drying_height_m"] = best
    save_mooring(m)

    log_activity(
        event_type="calibration_apply",
        message=(
            f"Drying height applied from observations: "
            f"{previous:.2f}m -> {best:.2f}m ({cal.get('confidence')})"
        ),
        severity="success",
        scope="mooring",
        mooring_id=mooring_id,
        details={
            "field": "drying_height_m",
            "previous": previous,
            "new": best,
            "confidence": cal.get("confidence"),
            "lower_bound": cal.get("lower_bound"),
            "upper_bound": cal.get("upper_bound"),
            "afloat_count": cal.get("afloat_count"),
            "aground_count": cal.get("aground_count"),
            # v2.10 sounding audit trail: enough to reconcile a
            # sounding-driven recalibration against the bounds after the
            # fact (validation 1 in CALIBRATION_NOTES).
            "sounding_count": cal.get("sounding_count"),
            "sounding_estimate": cal.get("sounding_estimate"),
            "sounding_sd": cal.get("sounding_sd"),
            "sounding_dryings": cal.get("sounding_dryings"),
            "floor_applied": cal.get("floor_applied"),
            "sounding_conflict": cal.get("sounding_conflict"),
            "height_source": "interpolate_height_at_time (pressure-blind)",
        },
    )

    _recompute_future_windows(mooring_id)

    return {
        "applied": True,
        "previous": previous,
        "new": best,
        "mooring": _public_mooring(get_mooring(mooring_id)),
        "calibration": _build_calibration_response(mooring_id),
    }


@app.post("/api/moorings/{mooring_id}/calibration/apply-wind-offset",
          dependencies=[Depends(require_mooring_pin)])
async def apply_wind_offset_calibration(mooring_id: int):
    """
    Apply the current shallow-side wind offset suggestion to the mooring's
    stored configuration (updates shallow_extra_depth_m), then recompute
    future access windows and regenerate the feed.
    """
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")

    cal = calibrate_wind_offset(mooring_id)
    suggested = cal.get("suggested_offset_m")
    if suggested is None:
        raise HTTPException(400, "No wind offset suggestion available to apply")

    try:
        previous = float(m.get("shallow_extra_depth_m") or 0.0)
    except (TypeError, ValueError):
        previous = 0.0

    if suggested <= previous + 0.01:
        return {
            "applied": False,
            "reason": (
                "current stored offset already meets or exceeds the "
                "observation-derived lower bound; no increase required"
            ),
            "previous": previous,
            "new": previous,
            "mooring": _public_mooring(m),
            "calibration": _build_calibration_response(mooring_id),
        }

    m["shallow_extra_depth_m"] = suggested
    save_mooring(m)

    log_activity(
        event_type="calibration_apply",
        message=(
            f"Wind offset applied from observations: "
            f"+{previous:.2f}m -> +{suggested:.2f}m ({cal.get('confidence')}, "
            f"{cal.get('observation_count')} observations)"
        ),
        severity="success",
        scope="mooring",
        mooring_id=mooring_id,
        details={
            "field": "shallow_extra_depth_m",
            "previous": previous,
            "new": suggested,
            "confidence": cal.get("confidence"),
            "observation_count": cal.get("observation_count"),
            "baseline_drying_height_m": cal.get("current_drying_height_m"),
        },
    )

    _recompute_future_windows(mooring_id)

    return {
        "applied": True,
        "previous": previous,
        "new": suggested,
        "mooring": _public_mooring(get_mooring(mooring_id)),
        "calibration": _build_calibration_response(mooring_id),
    }


# --- Wind jobs (debug / ops, PIN-gated) ---

@app.get("/api/moorings/{mooring_id}/wind-jobs",
         dependencies=[Depends(require_mooring_pin)])
async def list_wind_jobs(mooring_id: int):
    """
    PIN-gated: list the scheduled worst-case-grounding wind-sample jobs for
    this mooring. Visibility into the per-mooring wind scheduler for ops/debug.
    """
    from app.scheduler import scheduler

    prefix = f"wind_sample_m{mooring_id}_"
    jobs = []
    for j in scheduler.get_jobs():
        if (j.id or "").startswith(prefix):
            jobs.append({
                "job_id": j.id,
                "run_at": j.next_run_time.isoformat() if j.next_run_time else None,
                "next_hw": (j.kwargs or {}).get("next_hw_timestamp"),
            })
    jobs.sort(key=lambda x: x["run_at"] or "")
    return {"mooring_id": mooring_id, "count": len(jobs), "jobs": jobs}


@app.post("/api/moorings/{mooring_id}/wind-check/run",
          dependencies=[Depends(require_mooring_pin)])
async def run_wind_check_now(mooring_id: int):
    """
    PIN-gated: force an immediate wind check for this mooring against its next
    upcoming high water, exactly as the scheduled worst-case-grounding job
    would. Lets the wind-offset behaviour be verified without waiting for a
    real grounding. Returns the resulting stored event for that HW.
    """
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")
    if (not m.get("wind_offset_enabled")
            or not (m.get("shallow_direction") or "")
            or float(m.get("shallow_extra_depth_m") or 0.0) <= 0):
        raise HTTPException(400, "Wind offset is not configured for this mooring")

    from app.scheduler import wind_observation_job

    now = datetime.now(timezone.utc)
    events = get_ukho_tide_events(
        to_utc_str(now - timedelta(hours=1)),
        to_utc_str(now + timedelta(days=2)),
    )
    next_hw_ts = None
    for e in sorted(events, key=lambda ev: ev["timestamp"]):
        if e["event_type"] != "HighWater":
            continue
        hw_dt = dtparse.parse(e["timestamp"])
        if hw_dt.tzinfo is None:
            hw_dt = hw_dt.replace(tzinfo=timezone.utc)
        if hw_dt > now:
            next_hw_ts = to_utc_str(hw_dt)
            break
    if not next_hw_ts:
        raise HTTPException(400, "No upcoming high water in stored tide data")

    await wind_observation_job(mooring_id, next_hw_ts)

    stored = get_calendar_events(mooring_id, start=next_hw_ts, end=next_hw_ts)
    return {
        "ran": True,
        "mooring_id": mooring_id,
        "next_hw": next_hw_ts,
        "stored_event": stored[0] if stored else None,
    }


# --- UKHO Data ---

@app.post("/api/fetch-ukho")
async def trigger_ukho_fetch():
    """Manually trigger UKHO data fetch."""
    events, station_used = await fetch_tidal_events()
    if not events:
        log_activity(
            event_type="ukho_fetch",
            message="Manual UKHO fetch failed",
            severity="error",
        )
        raise HTTPException(502, "Failed to fetch UKHO data")

    station_label = "langstone" if station_used == UKHO_STATION_ID else "portsmouth"
    store_tide_events(events, source="ukho", station=station_label)

    fallback_note = f" (Portsmouth fallback, station {station_used})" if station_label == "portsmouth" else ""
    log_activity(
        event_type="ukho_fetch",
        message=f"Manual UKHO fetch: {len(events)} events stored{fallback_note}",
        severity="success" if station_label == "langstone" else "warning",
        details={
            "event_count": len(events),
            "trigger": "manual",
            "station_used": station_used,
            "station_label": station_label,
        },
    )

    # Reschedule per-mooring wind jobs from the freshly-stored tide data.
    # ensure_wind_jobs_scheduled reads tide events via get_ukho_tide_events,
    # which applies the Portsmouth->Langstone correction automatically, so the
    # manual-refresh path stays in step with the daily 02:00 path.
    ensure_wind_jobs_scheduled()

    # Regenerate the standalone Langstone tide feeds so manual refresh has
    # the same effect on those feeds as the daily 02:00 job. The harmonic
    # 180d feed regenerates from currently-stored harmonic predictions, so
    # it picks up the new UKHO data for days 0-7 even though the harmonic
    # set has not been recomputed by this manual trigger.
    try:
        generate_langstone_ukho_7d_feed()
        generate_langstone_harmonic_180d_feed()
        generate_langstone_ukho_7d_pressure_corrected_feed()
    except Exception as e:
        logger.warning(f"Langstone feed regeneration after manual fetch failed: {e}")

    return {"fetched": len(events), "source": "ukho", "station": station_label}


# --- Access Window Calculation ---

@app.post("/api/calculate")
async def calculate_access_windows(request: Request):
    """
    Calculate access windows for given parameters and data source.
    READ-ONLY: returns windows but does NOT store events or regenerate
    the feed. To push a calculation result into the mooring's stored
    events and iCal feed, call POST /api/moorings/{id}/feed/update
    afterwards (PIN-gated).

    Body:
        source: "ukho" | "harmonic"
        draught_m: float
        drying_height_m: float
        safety_margin_m: float
        mooring_id: int (optional - loads stored params when fields omitted)
        days: int (for harmonic, how far ahead)
        wind_offset_enabled: bool (optional - overrides stored mooring setting)
        shallow_direction: str (optional - overrides stored mooring setting)
        shallow_extra_depth_m: float (optional - overrides stored mooring setting)
    """
    data = await request.json()
    source = data.get("source", "ukho")
    mooring_id = data.get("mooring_id")

    mooring = None
    if mooring_id:
        mooring = get_mooring(int(mooring_id))

    draught = _coalesce(
        data.get("draught_m"),
        mooring["draught_m"] if mooring else None,
        default=1.0,
    )
    drying = _coalesce(
        data.get("drying_height_m"),
        mooring["drying_height_m"] if mooring else None,
        default=2.0,
    )
    margin = _coalesce(
        data.get("safety_margin_m"),
        mooring["safety_margin_m"] if mooring else None,
        default=0.3,
    )

    # Wind offset is not applied in this read-only preview. The scheduler is
    # the sole owner of wind adjustment: it applies the offset start-only to
    # the live feed at each vessel's worst-case grounding. /calculate (and the
    # /feed/update it feeds) therefore show and store *baseline* windows;
    # whatever wind is blowing at the next grounding adjusts the feed then.
    wind_offset = 0.0
    wind_info = None

    now = datetime.now(timezone.utc)
    # Look back 13 hours to ensure bracketing events for interpolation context.
    query_start = now - timedelta(hours=13)

    if source == "ukho":
        end = now + timedelta(days=7)
        # get_ukho_tide_events handles station preference and applies the
        # Portsmouth->Langstone offset transparently if only fallback data exists.
        events = get_ukho_tide_events(to_utc_str(query_start), to_utc_str(end))
        if not events:
            # Try fetching fresh from the API
            raw, station_used = await fetch_tidal_events()
            if raw:
                station_label = "langstone" if station_used == UKHO_STATION_ID else "portsmouth"
                store_tide_events(raw, source="ukho", station=station_label)
                # Re-query via get_ukho_tide_events so the 13h lookback window
                # is covered and the offset is applied if Portsmouth data was stored.
                events = get_ukho_tide_events(to_utc_str(query_start), to_utc_str(end))
        tide_source = "ukho"

    elif source == "harmonic":
        days = int(data.get("days", 30))
        end = now + timedelta(days=days)
        events = harmonic_predict_events(now, end)
        events = apply_offset(events)
        tide_source = "harmonic"

    else:
        raise HTTPException(400, f"Unknown source: {source}")

    if not events:
        return {"windows": [], "source": source, "event_count": 0, "message": "No tide data available"}

    # Barometric correction (v2.9): adjust predicted event heights for forecast
    # pressure before window computation. Gated on the system master AND this
    # mooring's opt-in, so the anonymous path (mooring is None) always stays
    # pure. Events are already Langstone-corrected here, so pressure is sampled
    # at the corrected event time. Per-event freshness is handled inside
    # apply_barometric_correction (events with no fresh forecast pass through
    # uncorrected). The reassigned list flows into both the vessel window
    # computation below and the tender pass, which share it.
    from app.config import get_barometric_enabled
    barometric_applied = bool(
        mooring and mooring.get("barometric_enabled") and get_barometric_enabled(False)
    )
    if barometric_applied:
        from app.barometric import apply_barometric_correction, make_pressure_provider
        events = apply_barometric_correction(events, make_pressure_provider())

    windows = compute_access_windows(
        events=events,
        draught_m=float(draught),
        drying_height_m=float(drying),
        safety_margin_m=float(margin),
        source=tide_source,
    )

    # Tender access pass. Uses the mooring's stored tender config; an
    # anonymous /calculate (no mooring_id) skips it.
    tender_windows, tender_depth = _compute_tender_windows(
        events, mooring, tide_source,
    )

    now_str = to_utc_str(now)
    windows = [w for w in windows if w["hw_timestamp"] >= now_str]

    # Merge tender data onto each main window so the UI can render a sub-row
    # without a second API call. Tender fields are also returned in
    # parameters for the round-trip to /feed/update.
    if tender_windows is not None:
        tender_by_hw = {tw["hw_timestamp"]: tw for tw in tender_windows}
        for w in windows:
            tw = tender_by_hw.get(w["hw_timestamp"])
            if tw is None:
                continue
            w["tender_start_time"] = tw.get("start_time")
            w["tender_end_time"] = tw.get("end_time")
            w["tender_always_accessible"] = bool(tw.get("always_accessible"))
            w["tender_below_threshold"] = bool(tw.get("below_threshold"))

    # Conservative 5-minute display rounding (v2.9, render-only). Attaches
    # display_* fields the UI renders, plus negligible_access when inward
    # rounding collapses a window. The raw start_time/end_time/tender_* are
    # left intact so this same response round-trips to /feed/update and
    # /export-ics at full precision (storage and the iCal emit-time rounding
    # both work from the raw edges).
    from app.window_display import display_fields
    for w in windows:
        w.update(display_fields(w))

    # v2: This endpoint is read-only. Storing events and regenerating the
    # iCal feed now happens via POST /api/moorings/{id}/feed/update, which
    # is PIN-gated. The UI passes this response body back to that endpoint
    # verbatim when the user clicks "Update Feed".

    parameters = {
        "draught_m": float(draught),
        "drying_height_m": float(drying),
        "safety_margin_m": float(margin),
        "wind_offset_m": wind_offset,
    }
    if tender_depth is not None:
        parameters["tender_min_depth_m"] = tender_depth
        parameters["tender_access_enabled"] = True

    return {
        "windows": windows,
        "source": source,
        "parameters": parameters,
        "wind_info": wind_info,
        "barometric_applied": barometric_applied,
        "event_count": len([w for w in windows if not w.get("below_threshold")]),
    }


# --- ICS Export ---

@app.post("/api/export-ics")
async def export_ics(request: Request):
    """Generate a downloadable .ics file from the last calculated windows."""
    data = await request.json()
    windows = data.get("windows", [])
    source = data.get("source", "ukho")
    boat_name = data.get("boat_name", "")
    mooring_id = data.get("mooring_id", 0)

    cal = None
    calc_params = None
    if mooring_id:
        m = get_mooring(int(mooring_id))
        if m:
            cal = calibrate_drying_height(int(mooring_id))
            calc_params = {
                "draught_m": m["draught_m"],
                "drying_height_m": m["drying_height_m"],
                "safety_margin_m": m["safety_margin_m"],
                "obs_calibrated": 1 if cal.get("confidence", "none") != "none" else 0,
            }
            if cal.get("confidence") == "none":
                cal = None

    ics_bytes = generate_export_ics(windows, source, boat_name, mooring_id, cal, calc_params)

    return Response(
        content=ics_bytes,
        media_type="text/calendar",
        headers={"Content-Disposition": "attachment; filename=tidal_access.ics"},
    )


# --- Feed Update (v2: decoupled from calculate) ---

@app.post("/api/moorings/{mooring_id}/feed/update",
          dependencies=[Depends(require_mooring_pin)])
async def update_mooring_feed(mooring_id: int, request: Request):
    """
    PIN-gated. Push a set of previously-calculated windows into the
    mooring's stored events, and regenerate the .ics feed file if the
    mooring has calendar_enabled. Separated from POST /api/calculate so
    that calculation can be a read-only operation: anyone can calculate
    windows or export a one-shot .ics, but only the PIN-holder can
    overwrite what the subscribed calendar feed serves.

    If calendar_enabled is false on the mooring, events are still stored
    (so a later re-enable picks them up) but the .ics file is not
    regenerated. The UI should hide the "Update Feed" button in this
    case; a direct call that bypasses the UI returns a feed_generated
    flag of False so the caller knows.

    Body (matches the shape returned by /api/calculate):
        source: str (ukho | harmonic)
        windows: [{...}, ...]
        parameters: {draught_m, drying_height_m, safety_margin_m, ...}
        wind_info: {direction, offset_m, shallow_side, applied} or null
    """
    data = await request.json()
    source = data.get("source", "")
    windows = data.get("windows", [])
    params = data.get("parameters", {}) or {}

    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")

    if not windows:
        raise HTTPException(400, "No windows supplied to update feed")
    if source not in ("ukho", "harmonic"):
        raise HTTPException(400, f"Invalid source: {source}")

    cal = calibrate_drying_height(mooring_id)
    calc_params = {
        "draught_m": params.get("draught_m", m["draught_m"]),
        "drying_height_m": params.get("drying_height_m", m["drying_height_m"]),
        "safety_margin_m": params.get("safety_margin_m", m["safety_margin_m"]),
        "obs_calibrated": 1 if cal.get("confidence", "none") != "none" else 0,
    }
    # Wind adjustment is owned by the scheduler (applied start-only to the live
    # feed at each grounding), not by this manual path. Any wind_info in the
    # request body is ignored so manual updates always store baseline windows.
    wind_details = None

    # Tender data round-trips on each window dict from /api/calculate.
    # Reconstruct the tender_windows list for store_windows_as_events.
    tender_windows = None
    tender_depth = (params or {}).get("tender_min_depth_m")
    if any(
        ("tender_start_time" in w) or w.get("tender_always_accessible")
        for w in windows
    ):
        tender_windows = [
            {
                "hw_timestamp": w["hw_timestamp"],
                "start_time": w.get("tender_start_time"),
                "end_time": w.get("tender_end_time"),
                "always_accessible": w.get("tender_always_accessible"),
                "below_threshold": w.get("tender_below_threshold", False),
            }
            for w in windows
        ]

    # Always store events - they are internal state, used by serve_feed
    # on every request, and preserved across calendar_enabled toggles.
    store_windows_as_events(
        windows, mooring_id, source, m.get("boat_name", ""),
        calc_params=calc_params, wind_details=wind_details,
        tender_windows=tender_windows,
        tender_min_depth_m=tender_depth,
    )

    feed_generated = False
    if m.get("calendar_enabled"):
        feed_cal = cal if cal.get("confidence") != "none" else None
        generate_feed_for_mooring(mooring_id, m.get("boat_name", ""), feed_cal)
        feed_generated = True

    log_activity(
        event_type="feed_generation",
        message=(
            f"Calendar feed updated ({len(windows)} windows from {source.upper()})"
            if feed_generated
            else f"Events stored but feed file not regenerated - calendar disabled for mooring #{mooring_id}"
        ),
        severity="info" if feed_generated else "warning",
        scope="mooring",
        mooring_id=mooring_id,
        details={
            "window_count": len(windows),
            "source": source,
            "trigger": "user_update_feed",
            "feed_generated": feed_generated,
        },
    )

    return {
        "status": "ok",
        "window_count": len(windows),
        "source": source,
        "feed_generated": feed_generated,
    }


# --- Activity Log ---

@app.get("/api/activity-log")
async def get_activity(
    scope: str = None,
    mooring_id: int = None,
    event_type: str = None,
    severity: str = None,
    limit: int = 500,
):
    """Query the activity log. Filter by scope (system/mooring), mooring_id, event_type, severity."""
    return get_activity_log(
        scope=scope,
        mooring_id=mooring_id,
        event_type=event_type,
        severity=severity,
        limit=min(limit, 1000),
    )


# --- iCal Feed ---

@app.get("/feeds/mooring_{mooring_id:int}.ics")
async def serve_feed(mooring_id: int):
    """Serve the subscribable iCal feed for a mooring.
    Always regenerates from DB to ensure freshness."""
    m = get_mooring(mooring_id)
    if not m:
        raise HTTPException(404, "Mooring not found")

    cal = calibrate_drying_height(mooring_id)
    if cal.get("confidence") == "none":
        cal = None
    feed_path = generate_feed_for_mooring(mooring_id, m.get("boat_name", ""), cal)

    return Response(
        content=feed_path.read_bytes(),
        media_type="text/calendar",
        headers={
            "Content-Type": "text/calendar; charset=utf-8",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


# --- Model Config ---

@app.get("/api/model-config")
async def get_model_config():
    """
    Get the current model configuration.

    Returns the bundled model configuration shipped with the application.
    Read-only as of v2.5.5; the previous POST endpoint was removed because
    it wrote to a volume-persisted operative copy that the cached loader
    never picked up - silently lossy. The configuration now lives only in
    app/model_config.json in the repo and reaches running containers via
    rebuild + restart.
    """
    return load_model_config()


# --- Tide Data ---

@app.get("/api/tide-data")
async def get_stored_tide_data(start: str = None, end: str = None):
    """Get stored tide data within a date range."""
    if not start:
        start = to_utc_str(datetime.now(timezone.utc))
    if not end:
        end = to_utc_str(datetime.now(timezone.utc) + timedelta(days=7))
    return get_tide_events(start, end)


@app.get("/api/tides")
async def get_tides(range: str = "forecast"):
    """
    Return tide events for the Tides tab.

    range='forecast' (default): UKHO events from now to now+7 days.
    range='extended':            UKHO events for days 0-7 plus harmonic
                                  predictions for days 7-180. Each event
                                  carries an extra `data_source` field set
                                  to 'ukho' or 'harmonic' so the UI can
                                  show a per-row badge and insert a
                                  visible break between sources.
    range='history':              UKHO events from now-365 days up to now.

    Each event is returned with its `station` field intact ('langstone' or
    'portsmouth') so the UI can show a per-row source badge. The Portsmouth
    secondary-port offset is NOT applied here for stored UKHO events - the
    UI shows raw stored values from each station, with a badge indicating
    which one. Harmonic events have already been Langstone-corrected at
    storage time (scheduler applies the offset before write), so they need
    no further correction here; their station field is set to 'langstone'
    for consistency.

    Note: this endpoint is intentionally ungated - tide data is public.
    """
    if range not in ("forecast", "extended", "history"):
        raise HTTPException(400, "range must be 'forecast', 'extended', or 'history'")

    now = datetime.now(timezone.utc)
    if range == "forecast":
        start = now
        end = now + timedelta(days=7)
        events = get_tide_events(
            to_utc_str(start), to_utc_str(end), source="ukho"
        )
        for ev in events:
            ev["data_source"] = "ukho"
        return {"range": range, "events": events, "count": len(events)}

    if range == "history":
        # Rolling 12-month window. The cleanup_old_tide_data scheduled job
        # prevents data older than this from accumulating, so the upper
        # limit of the query is effectively bounded by retention.
        start = now - timedelta(days=365)
        end = now
        events = get_tide_events(
            to_utc_str(start), to_utc_str(end), source="ukho"
        )
        for ev in events:
            ev["data_source"] = "ukho"
        return {"range": range, "events": events, "count": len(events)}

    # range == 'extended': UKHO days 0-7 + harmonic days 7-180.
    ukho_end = now + timedelta(days=7)
    harmonic_end = now + timedelta(days=180)

    ukho_events = get_tide_events(
        to_utc_str(now), to_utc_str(ukho_end), source="ukho"
    )
    for ev in ukho_events:
        ev["data_source"] = "ukho"

    # Harmonic predictions are already Langstone-corrected (scheduler applies
    # the offset before storing). They include neither station nor source
    # columns from get_tide_events shape; tag them so the UI sees a uniform
    # event dict.
    harmonic_start = now + timedelta(days=7)
    harmonic_rows = get_harmonic_predictions(
        to_utc_str(harmonic_start), to_utc_str(harmonic_end), latest_only=True
    )
    harmonic_events = []
    for h in harmonic_rows:
        harmonic_events.append({
            "timestamp": h["timestamp"],
            "height_m": h["height_m"],
            "event_type": h["event_type"],
            "station": "langstone",  # post-offset; for badge consistency
            "source": "harmonic",
            "data_source": "harmonic",
            "is_approximate_time": True,
            "is_approximate_height": True,
        })

    # Defensive deduplication: drop any harmonic event whose timestamp is
    # within 90 minutes of a UKHO event of the same type. UKHO wins. Same
    # rule as in the 180d feed generator.
    ukho_index = []
    for ue in ukho_events:
        ue_dt = dtparse.parse(ue["timestamp"])
        if ue_dt.tzinfo is None:
            ue_dt = ue_dt.replace(tzinfo=timezone.utc)
        ukho_index.append((ue_dt, ue["event_type"]))
    filtered_harmonic = []
    for he in harmonic_events:
        he_dt = dtparse.parse(he["timestamp"])
        if he_dt.tzinfo is None:
            he_dt = he_dt.replace(tzinfo=timezone.utc)
        clash = any(
            ukho_et == he["event_type"]
            and abs((he_dt - ukho_dt).total_seconds()) < 5400
            for ukho_dt, ukho_et in ukho_index
        )
        if not clash:
            filtered_harmonic.append(he)

    combined = ukho_events + filtered_harmonic
    combined.sort(key=lambda e: e["timestamp"])
    return {
        "range": range,
        "events": combined,
        "count": len(combined),
        "ukho_count": len(ukho_events),
        "harmonic_count": len(filtered_harmonic),
    }


# --- Tidal curve (v2.8 / v2.8.1) ---
#
# Sampled height-vs-time series for the requested local day, used by the
# Tidal Curve panel in the UI. Date-based source switch (v2.8.1):
#   - days_ahead in [..., 6]: UKHO. If today's date and bracketing UKHO
#     events are missing, a one-shot lazy UKHO fetch is triggered
#     (subject to a 30s cooldown to bound retries on outages).
#   - days_ahead >= 7: harmonic. Stored harmonic_predictions are queried
#     via get_harmonic_predictions(latest_only=True) and piped through
#     the same _interpolate_from_parsed curve model so the displayed
#     curve and any access windows computed elsewhere stay consistent.
#     No on-demand generation; the scheduler keeps 180 days populated
#     at 02:00 and a startup warm-up covers the first-run gap.
#
# The cutoff is fixed at day 7 to avoid mixed-source curves on the same
# date (e.g. UKHO HW + harmonic LW from a partial day) - the same
# rationale used by /api/tides?range=extended.

_tide_curve_lazy_fetch_last_attempt: Optional[datetime] = None
_TIDE_CURVE_LAZY_FETCH_COOLDOWN_SECONDS = 30

# v2.8.1: harmonic switchover offset in days. Set to 7 because the UKHO
# Discovery tier covers the next 7 calendar days.
_HARMONIC_SWITCHOVER_DAYS = 7

# v2.8.1: forward extent of the date picker in days. Bounded by the
# 180-day harmonic horizon. The picker greys out at this limit.
_AVAILABLE_RANGE_FORWARD_DAYS = 180


def _available_range(reference_local_date) -> dict:
    """
    Return the inclusive local-date range (as YYYY-MM-DD strings) over
    which the UI may select dates for the tide curve.

    Bounds:
      - earliest: the earliest UKHO timestamp stored in tide_data. The
                  curve uses UKHO for dates in the past where data exists.
                  Harmonic predictions are forward-only in this app's
                  storage workflow, so history is not extended into the
                  harmonic horizon.
      - latest:   today + 180. This is the harmonic horizon; the
                  scheduler keeps the next 180 days populated.

    Falls back to reference_local_date on either bound if a DB error occurs.
    """
    try:
        import pytz
        from app.database import db_connection
        tz = pytz.timezone(DEFAULT_TIMEZONE)
        with db_connection() as conn:
            row = conn.execute(
                "SELECT MIN(timestamp), MAX(timestamp) FROM tide_data WHERE source = 'ukho'"
            ).fetchone()
        earliest_iso = row[0] if row else None
        if earliest_iso:
            earliest_local = dtparse.parse(earliest_iso).astimezone(tz).date()
        else:
            earliest_local = reference_local_date
        latest_local = reference_local_date + timedelta(days=_AVAILABLE_RANGE_FORWARD_DAYS)
        return {
            "earliest": earliest_local.isoformat(),
            "latest": latest_local.isoformat(),
        }
    except Exception as e:
        logger.warning(f"available_range query failed: {e}")
        return {"earliest": reference_local_date.isoformat(),
                "latest": reference_local_date.isoformat()}


@app.get("/api/tide-curve")
async def get_tide_curve(date: Optional[str] = None, step_minutes: int = 5):
    """
    Return a sampled tidal height curve for the requested local day.

    Query params:
      date: ISO date YYYY-MM-DD in Europe/London. Default = today (local).
      step_minutes: sample interval in minutes (default 5, min 1, max 60).

    Returns:
      {
        "date": "YYYY-MM-DD",
        "step_minutes": int,
        "samples": [{"time": "...Z", "height_m": float}, ...],
        "events_in_window": [{"type": "HighWater"|"LowWater", "time": "...Z", "height_m": float}, ...],
        "source_used": "ukho" | "harmonic" | null,
        "reason": "no_ukho_data" | "no_harmonic_data" | None,
        "lazy_fetch_attempted": bool,
        "available_range": {"earliest": "YYYY-MM-DD", "latest": "YYYY-MM-DD"},
        "spring_neap": "spring"|"mid"|"neap"|None,
        "sunrise": "...Z" | None,
        "sunset": "...Z" | None,
      }

    Source switch (v2.8.1): today+0..6 uses UKHO (with lazy fetch on
    today's date if missing); today+7 onwards uses stored harmonic
    predictions. The cutoff is fixed by date, not by data availability,
    to avoid mixed-source curves on a single day.
    """
    global _tide_curve_lazy_fetch_last_attempt

    if step_minutes < 1 or step_minutes > 60:
        raise HTTPException(400, "step_minutes must be between 1 and 60")

    import pytz
    tz = pytz.timezone(DEFAULT_TIMEZONE)
    now_utc = datetime.now(timezone.utc)

    if date:
        try:
            local_date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "date must be YYYY-MM-DD")
    else:
        local_date = now_utc.astimezone(tz).date()

    # Local-day window: 00:00 inclusive to 24:00 inclusive (sample at both endpoints).
    local_midnight = tz.localize(datetime.combine(local_date, datetime.min.time()))
    local_end = local_midnight + timedelta(days=1)
    day_start_utc = local_midnight.astimezone(timezone.utc)
    day_end_utc = local_end.astimezone(timezone.utc)

    # Query a wider window so we have bracketing HW/LW events. Spring tides
    # at Langstone are ~6h12m apart, so ±18h is comfortably more than one
    # cycle either side and guarantees bracketing if data exists.
    query_start = to_utc_str(day_start_utc - timedelta(hours=18))
    query_end = to_utc_str(day_end_utc + timedelta(hours=18))

    def _bracketed(events: list[dict]) -> bool:
        """True iff events contain at least one entry before day_start_utc
        and at least one entry after day_end_utc."""
        has_before = any(dtparse.parse(e["timestamp"]) <= day_start_utc for e in events)
        has_after = any(dtparse.parse(e["timestamp"]) >= day_end_utc for e in events)
        return has_before and has_after

    today_local = now_utc.astimezone(tz).date()
    days_ahead = (local_date - today_local).days
    use_harmonic = days_ahead >= _HARMONIC_SWITCHOVER_DAYS
    is_today = (local_date == today_local)
    lazy_fetch_attempted = False

    if use_harmonic:
        # Harmonic branch (day 7+). No lazy generation - the scheduler
        # and the startup warm-up are responsible for keeping the table
        # populated. If the cycle is missing, return a no_harmonic_data
        # response and let the UI prompt the operator to regen.
        from app.database import get_harmonic_predictions
        events = get_harmonic_predictions(query_start, query_end, latest_only=True)
        source_used = "harmonic"
        no_data_reason = "no_harmonic_data"
    else:
        # UKHO branch (day -inf..6). Lazy fetch only fires on today's date.
        events = get_ukho_tide_events(query_start, query_end)
        source_used = "ukho"
        no_data_reason = "no_ukho_data"

        if not _bracketed(events) and is_today:
            now = datetime.now(timezone.utc)
            last = _tide_curve_lazy_fetch_last_attempt
            cooldown_active = (
                last is not None
                and (now - last).total_seconds() < _TIDE_CURVE_LAZY_FETCH_COOLDOWN_SECONDS
            )
            if not cooldown_active:
                _tide_curve_lazy_fetch_last_attempt = now
                lazy_fetch_attempted = True
                try:
                    fetched, station_used = await fetch_tidal_events()
                    if fetched:
                        station_label = "langstone" if station_used == UKHO_STATION_ID else "portsmouth"
                        store_tide_events(fetched, source="ukho", station=station_label)
                        log_activity(
                            event_type="ukho_fetch",
                            message=f"Tide curve lazy fetch: {len(fetched)} events stored",
                            severity="info",
                            details={
                                "event_count": len(fetched),
                                "trigger": "tide_curve_lazy_fetch",
                                "station_used": station_used,
                                "station_label": station_label,
                                "requested_date": local_date.isoformat(),
                            },
                        )
                        events = get_ukho_tide_events(query_start, query_end)
                except Exception as e:
                    logger.warning(f"Tide curve lazy UKHO fetch failed: {e}")

    # Parse + sort once for the inner sampling loop. We do NOT gate on
    # strict bracketing here (events both before AND after the day): at
    # the UKHO horizon the trailing bracket may be missing yet most of
    # the day is still interpolable. Sample what we can and only
    # short-circuit if the result is empty.
    parsed = []
    for ev in events:
        dt = dtparse.parse(ev["timestamp"])
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        parsed.append((dt, ev["height_m"], ev["event_type"]))
    parsed.sort(key=lambda x: x[0])

    step = timedelta(minutes=step_minutes)
    samples = []
    t = day_start_utc
    while t <= day_end_utc:
        h = _interpolate_from_parsed(t, parsed)
        if h is not None:
            samples.append({
                "time": to_utc_str(t),
                "height_m": round(h, 3),
            })
        t += step

    if not samples:
        sun_times_fallback = _compute_sun_times(local_date)
        return {
            "date": local_date.isoformat(),
            "step_minutes": step_minutes,
            "samples": [],
            "events_in_window": [],
            "source_used": None,
            "reason": no_data_reason,
            "lazy_fetch_attempted": lazy_fetch_attempted,
            "spring_neap": None,
            "sunrise": sun_times_fallback["sunrise"],
            "sunset": sun_times_fallback["sunset"],
            "available_range": _available_range(today_local),
        }

    # Events strictly inside the day window (for HW/LW markers).
    events_in_window = [
        {
            "type": et,
            "time": to_utc_str(dt),
            "height_m": round(h, 3),
        }
        for (dt, h, et) in parsed
        if day_start_utc <= dt <= day_end_utc
    ]

    # Spring/Neap/Mid classification for the displayed date (v2.8).
    try:
        from app.tide_state import classify_spring_neap
        spring_neap = classify_spring_neap(local_date)
    except Exception as e:
        logger.warning(f"Spring/neap classification skipped: {e}")
        spring_neap = None

    sun_times = _compute_sun_times(local_date)

    return {
        "date": local_date.isoformat(),
        "step_minutes": step_minutes,
        "samples": samples,
        "events_in_window": events_in_window,
        "source_used": source_used,
        "reason": None,
        "lazy_fetch_attempted": lazy_fetch_attempted,
        "spring_neap": spring_neap,
        "sunrise": sun_times["sunrise"],
        "sunset": sun_times["sunset"],
        "available_range": _available_range(today_local),
    }


def _compute_sun_times(local_date) -> dict:
    """
    Astronomical sunrise/sunset for the configured lat/lon on a local
    date. Returns ISO-Z UTC strings, or None for either if astral fails.
    Adding ~30 KB of pip-installed code; the computation is deterministic
    and doesn't need an external API call.
    """
    try:
        import pytz
        from astral import LocationInfo
        from astral.sun import sun
        loc = LocationInfo(
            name="Langstone", region="UK", timezone=DEFAULT_TIMEZONE,
            latitude=LOCATION_LAT, longitude=LOCATION_LON,
        )
        s = sun(loc.observer, date=local_date, tzinfo=pytz.timezone(DEFAULT_TIMEZONE))
        return {
            "sunrise": to_utc_str(s["sunrise"]),
            "sunset": to_utc_str(s["sunset"]),
        }
    except Exception as e:
        logger.warning(f"Sun-times computation failed: {e}")
        return {"sunrise": None, "sunset": None}


# --- Standalone Langstone tide feeds ---
#
# These two feeds are unaffiliated with any mooring and serve UKHO tide
# events (next 7 days) and a UKHO+harmonic merge (next 180 days). They are
# regenerated daily by the scheduler at 02:00. Both routes regenerate the
# file on demand if it is missing or stale - this covers first deployment
# (no scheduler run yet) and any in-day regeneration triggered by a manual
# UKHO refresh.

@app.get("/feeds/Langstone_UKHO_7d.ics")
async def serve_langstone_ukho_7d():
    """Serve the Langstone UKHO 7-day tide feed. Regenerates if missing."""
    feed_path = FEEDS_DIR / "Langstone_UKHO_7d.ics"
    if not feed_path.exists():
        feed_path = generate_langstone_ukho_7d_feed()
    return Response(
        content=feed_path.read_bytes(),
        media_type="text/calendar",
        headers={
            "Content-Type": "text/calendar; charset=utf-8",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/feeds/Langstone_UKHO_7d_PressureCorrected.ics")
async def serve_langstone_ukho_7d_pressure_corrected():
    """Serve the barometric pressure-corrected UKHO 7-day tide feed.
    Regenerates if missing."""
    feed_path = FEEDS_DIR / "Langstone_UKHO_7d_PressureCorrected.ics"
    if not feed_path.exists():
        feed_path = generate_langstone_ukho_7d_pressure_corrected_feed()
    return Response(
        content=feed_path.read_bytes(),
        media_type="text/calendar",
        headers={
            "Content-Type": "text/calendar; charset=utf-8",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@app.get("/feeds/Langstone_Harmonic_180d.ics")
async def serve_langstone_harmonic_180d():
    """Serve the Langstone 180-day tide feed. Regenerates if missing."""
    feed_path = FEEDS_DIR / "Langstone_Harmonic_180d.ics"
    if not feed_path.exists():
        feed_path = generate_langstone_harmonic_180d_feed()
    return Response(
        content=feed_path.read_bytes(),
        media_type="text/calendar",
        headers={
            "Content-Type": "text/calendar; charset=utf-8",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


# --- Current Conditions ---

@app.get("/api/conditions")
async def current_conditions():
    """
    Return combined tide and weather conditions at Langstone Harbour.

    Cached for 15 minutes. The scheduler refresh job keeps the cache
    warm and the pressure history populated even when no user is
    viewing the page.

    Ungated (no PIN required) - weather and tide state are public.
    """
    result = await get_current_conditions()
    return result


# --- Wind ---

@app.get("/api/wind/current")
async def get_current_wind():
    """Fetch current wind conditions."""
    data = await fetch_current_wind()
    if not data:
        raise HTTPException(502, "Could not fetch wind data")
    return data


# --- Calendar Events ---

@app.get("/api/moorings/{mooring_id}/events")
async def get_mooring_events(mooring_id: int, start: str = None, end: str = None):
    """Get calendar events for a mooring."""
    return get_calendar_events(mooring_id, start, end)


# --- Config Status ---

@app.get("/api/config/status")
async def config_status():
    """Return which optional features are available based on env config."""
    owm_configured = bool(OWM_API_KEY) and OWM_API_KEY not in (
        "", "your_openweathermap_api_key_here"
    )
    return {
        "owm_available": owm_configured,
    }
